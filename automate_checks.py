"""
Enhanced automate_checks.py with:
- Automatic cookie refresh every 20 minutes
- File upload support
- Better Excel data extraction
"""

import json
import time
import logging
import sys
import threading
import shutil
import base64
import uuid
import subprocess
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import load_workbook

import config
from ocr_solver import CaptchaSolver
from gemini_ocr import GeminiSolver
from progress_db import ProgressTracker


# Setup logging
def setup_logging():
    """Configure logging to file and console"""
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    
    logger = logging.getLogger()
    logger.setLevel(getattr(logging, config.LOG_LEVEL))
    
    file_handler = logging.FileHandler(config.LOG_FILE, encoding='utf-8')
    file_handler.setFormatter(logging.Formatter(log_format))
    logger.addHandler(file_handler)
    
    if config.CONSOLE_OUTPUT:
        # Use stderr with error handling to avoid Unicode issues on Windows console
        console_handler = logging.StreamHandler(sys.stderr)
        console_handler.setFormatter(logging.Formatter(log_format))
        logger.addHandler(console_handler)
    
    return logger


logger = setup_logging()


class CookieRefresher:
    """Handles automatic cookie refresh"""
    
    def __init__(self, callback, interval_minutes=20):
        self.callback = callback
        self.interval = interval_minutes * 60  # Convert to seconds
        self.stop_flag = threading.Event()
        self.thread = None
    
    def start(self):
        """Start the cookie refresh thread"""
        self.thread = threading.Thread(target=self._refresh_loop, daemon=True)
        self.thread.start()
        logger.info(f"Cookie refresher started (interval: {self.interval/60} minutes)")
    
    def stop(self):
        """Stop the cookie refresh thread"""
        self.stop_flag.set()
        if self.thread:
            self.thread.join(timeout=5)
        logger.info("Cookie refresher stopped")
    
    def _refresh_loop(self):
        """Background thread that refreshes cookies periodically"""
        while not self.stop_flag.is_set():
            time.sleep(self.interval)
            if not self.stop_flag.is_set():
                try:
                    self.callback()
                except Exception as e:
                    logger.error(f"Error refreshing cookies: {e}")


class CheckAutomation:
    """Main automation class for check editing"""
    
    def __init__(self):
        """Initialize automation"""
        self.session = requests.Session()
        
        # Configure retry adapter with connection pooling for better performance
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry
        
        retry_strategy = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "OPTIONS", "POST"]
        )
        
        # Enhanced connection pooling configuration
        pool_size = getattr(config, 'CONNECTION_POOL_SIZE', 50)
        pool_maxsize = getattr(config, 'CONNECTION_POOL_MAXSIZE', 50)
        pool_block = getattr(config, 'CONNECTION_POOL_BLOCK', False)
        
        adapter = HTTPAdapter(
            max_retries=retry_strategy,
            pool_connections=pool_size,  # Number of connection pools to cache
            pool_maxsize=pool_maxsize,    # Max connections per pool
            pool_block=pool_block          # Don't block when pool is exhausted
        )
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)
        
        logger.info(f"HTTP connection pool configured: size={pool_size}, maxsize={pool_maxsize}")
        
        self.session.headers.update(config.HEADERS)
        
        # Set cookies if provided
        if config.SESSION_COOKIES:
            self.session.cookies.update(config.SESSION_COOKIES)
            logger.info("Session cookies loaded")
        
        # Initialize cookie refresher
        self.cookie_refresher = CookieRefresher(self.refresh_cookies, interval_minutes=20)
        
        # Initialize OCR solver
        self.gemini_solver = None
        self.captcha_solver = None
        
        if config.OCR_ENGINE == 'gemini':
            self.gemini_solver = GeminiSolver()
            logger.info("Gemini Batch OCR initialized")
        else:
            # Only initialize CaptchaSolver for non-Gemini engines
            if config.OCR_ENGINE == 'easyocr':
                ocr_config = {'langs': config.EASYOCR_LANGS, 'gpu': config.EASYOCR_GPU}
            elif config.OCR_ENGINE == 'paddle':
                ocr_config = {'lang': 'en', 'use_angle_cls': True}
            else:
                ocr_config = {'cmd': config.TESSERACT_CMD, 'config': config.TESSERACT_CONFIG}
            
            self.captcha_solver = CaptchaSolver(
                engine=config.OCR_ENGINE,
                config=ocr_config
            )
            logger.info(f"OCR engine initialized: {config.OCR_ENGINE}")
        
        # Progress tracking (SQLite-based)
        self.progress_tracker = ProgressTracker()
        self.progress = self.progress_tracker.get_all()
        logger.info(f"Loaded progress: {len(self.progress)} completed checks")
        self.stats = {
            'total': 0,
            'processed': 0,
            'successful': 0,
            'failed': 0,
            'captcha_failures': 0,
            'start_time': None,
            'end_time': None,
            'last_cookie_refresh': datetime.now()
        }
        
        # Load ZIP file if needed
        self.zip_file_content = None
        if Path(config.ZIP_FILE).exists():
            with open(config.ZIP_FILE, 'rb') as f:
                self.zip_file_content = f.read()
            logger.info(f"Loaded ZIP file: {config.ZIP_FILE} ({len(self.zip_file_content)} bytes)")
        
        # Session refresh tracking
        self.session_expired = False
        self.unauthorized_count = 0
    
    def refresh_cookies(self):
        """
        Refresh session cookies
        
        This method can be enhanced to automatically re-login if needed.
        For now, it logs that cookies should be refreshed manually.
        """
        logger.info("Cookie refresh triggered - Please update SESSION_COOKIES in config.py if needed")
        self.stats['last_cookie_refresh'] = datetime.now()
        
        # Automatic session refresh is now handled by handle_session_expired()
        # For example:
        # if config.USERNAME and config.PASSWORD:
        #     self.login(config.USERNAME, config.PASSWORD)

    def _get_bearer_token(self):
        """Retrieve bearer token from config, session cookies, or session headers.

        Returns a string starting with 'Bearer ' if a token is found, otherwise an empty string.
        """
        # Try config first
        token = ''
        try:
            if isinstance(config.SESSION_COOKIES, dict):
                token = config.SESSION_COOKIES.get('bearer_token') or config.SESSION_COOKIES.get('Authorization') or ''
        except Exception:
            token = ''

        # Then try session cookies
        if not token:
            try:
                token = self.session.cookies.get('bearer_token') or self.session.cookies.get('Authorization') or ''
            except Exception:
                token = ''

        # Then try session headers
        if not token:
            token = self.session.headers.get('Authorization', '') or ''

        if token and not token.startswith('Bearer '):
            token = f'Bearer {token}'

        return token
    
    def upload_file(self, check_number):
        """
        Upload ZIP file for a check using base64 JSON upload
        
        The API uses JSON (not multipart) with base64-encoded file content.
        Requires Bearer token authentication.
        
        Args:
            check_number (str): Check number
            
        Returns:
            str: Uploaded file identifier or empty string
        """
        if not self.zip_file_content:
            logger.warning("No ZIP file content loaded. Skipping upload.")
            return ""
        
        # Retry loop for file upload
        max_retries = getattr(config, 'UPLOAD_MAX_RETRIES', 3)
        retry_delay = getattr(config, 'UPLOAD_RETRY_DELAY', 5)
        
        for attempt in range(max_retries):
            try:
                # Small random delay before upload to prevent rate limiting
                # Small random delay before upload to prevent rate limiting
                # REMOVED: Artificial delay removed for performance
                # import random
                # if attempt > 0:
                #     sleep_time = retry_delay * (attempt + 1) + random.uniform(0.5, 2.0)
                #     logger.info(f"Retry {attempt+1}/{max_retries} for upload in {sleep_time:.1f}s...")
                #     time.sleep(sleep_time)
                # else:
                #     # time.sleep(random.uniform(0.5, 1.5))
                #     pass
                
                upload_start = time.time()
                logger.info(f"Uploading file for check {check_number} (Attempt {attempt+1})...")
                
                # Encode file to base64
                file_base64 = base64.b64encode(self.zip_file_content).decode('utf-8')
                
                # Generate dynamic repository ID (mimic browser behavior)
                repository_id = uuid.uuid4().hex
                
                # Build JSON payload matching the successful HAR file request exactly
                payload = {
                    "lang": "ru",
                    "docType": "application/x-zip-compressed",
                    "pinCode": config.DEFAULT_PIN_CODE,
                    "repositoryId": repository_id,  # Dynamic ID
                    "docDate": datetime.now().strftime("%d.%m.%Y"),
                    "interactiveId": config.INTERACTIVE_ID,
                    "tin": config.DEFAULT_TIN,
                    "docNum": "docNum",
                    "fileName": "АСИЛБЕКОВА.zip",
                    "contentType": "application/x-zip-compressed",
                    "file": file_base64
                }
                
                # Serialize payload manually
                payload_str = json.dumps(payload)
                
                # Log payload size for debugging
                payload_size_mb = len(payload_str) / (1024 * 1024)
                logger.debug(f"Payload size: {payload_size_mb:.2f} MB")
                
                # Prepare headers with Bearer token if available
                headers = dict(self.session.headers)
                bearer_token = self._get_bearer_token()

                if bearer_token:
                    headers['Authorization'] = bearer_token
                    logger.debug(f"Using Bearer token: {bearer_token[:20]}...")
                else:
                    logger.warning("⚠️ No Bearer token configured - upload may fail!")
                    logger.warning("   Add 'bearer_token' to SESSION_COOKIES in config.py or set Authorization header")
                
                # Make the request
                response = self.session.post(
                    config.FILE_UPLOAD_URL,
                    data=payload_str,
                    headers=headers,
                    timeout=config.REQUEST_TIMEOUT
                )
                
                if response.status_code == 200:
                    try:
                        result = response.json() if response.text else {}
                        if result.get('success') or result.get('fileGuid'):
                            file_id = result.get('fileGuid') or result.get('id') or ''
                            upload_duration = time.time() - upload_start
                            logger.info(f"File uploaded successfully: {file_id} (took {upload_duration:.1f}s)")
                            return file_id
                        else:
                            logger.warning(f"Upload returned 200 but no file ID in response: {result}")
                            # Don't retry immediately if it's a logic error, but here we might want to
                            # For now, treat as failure and retry
                    except Exception as json_err:
                        logger.error(f"Failed to parse upload response: {json_err}")
                elif response.status_code == 404:
                    # Try fallback upload URL
                    logger.info(f"Got 404, trying fallback upload URL")
                    alt_url = config.BASE_URL + "/file/repository-set-file"
                    try:
                        response2 = self.session.post(
                            alt_url,
                            data=payload_str,
                            headers=headers,
                            timeout=config.REQUEST_TIMEOUT
                        )
                        
                        if response2.status_code == 200:
                            result = response2.json() if response2.text else {}
                            file_id = result.get('fileGuid') or result.get('id') or ''
                            if file_id:
                                logger.info(f"File uploaded via fallback URL: {file_id}")
                                return file_id
                    except Exception as e:
                        logger.warning(f"Fallback upload also failed: {e}")
                elif response.status_code == 401:
                    logger.error("❌ 401 Unauthorized during file upload")
                    return 'unauthorized'
                else:
                    logger.error(f"Upload failed with status {response.status_code}: {response.text[:200]}")
            
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                logger.warning(f"Timeout/Connection error during upload (Attempt {attempt+1}): {e}")
                # Continue to next retry
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 401:
                    logger.error("❌ 401 Unauthorized during file upload")
                    return 'unauthorized'
                logger.error(f"HTTP error uploading file: {e}")
            except Exception as e:
                logger.error(f"Error uploading file: {e}", exc_info=True)
                # For unknown errors, maybe don't retry? Or retry cautiously.
                # Let's retry for now.
        
        logger.error(f"Failed to upload file after {max_retries} attempts")
        return ""
    
    def delete_failed_captcha_images(self, check_number):
        """
        Delete any saved failed CAPTCHA images for a specific check number
        
        Args:
            check_number (str): Check number to delete debug images for
        """
        try:
            pattern = f"failed_captcha_{check_number}_*.png"
            deleted_count = 0
            for p in Path(".").glob(pattern):
                try:
                    p.unlink()
                    deleted_count += 1
                except Exception as e:
                    logger.debug(f"Error deleting {p}: {e}")
            
            if deleted_count > 0:
                logger.debug(f"Deleted {deleted_count} failed CAPTCHA image(s) for check {check_number}")
        except Exception as e:
            logger.debug(f"Error cleaning up failed CAPTCHA images: {e}")
    
    def load_progress(self):
        """Load progress from SQLite database"""
        # This method is now handled by ProgressTracker in __init__
        # Kept for backward compatibility
        return self.progress_tracker.get_all()
    
    def save_progress(self, new_checks=None):
        """Save progress to SQLite database (fast, corruption-proof)"""
        try:
            if new_checks:
                # Incremental save - very fast!
                self.progress_tracker.add_batch(new_checks)
            else:
                # Full save (fallback)
                self.progress_tracker.add_batch(self.progress)
            # SQLite only - no JSON export for maximum speed
        except Exception as e:
            logger.error(f"Error saving progress: {e}")
    
    def get_captcha(self):
        """
        Fetch CAPTCHA from server
        
        Returns:
            tuple: (captcha_id, captcha_image_base64) or (None, None) on error
        """
        try:
            logger.debug("Fetching CAPTCHA...")
            
            # Prepare headers (Bearer token may be needed)
            headers = dict(self.session.headers)
            bearer_token = self._get_bearer_token()

            if bearer_token:
                headers['Authorization'] = bearer_token

            response = self.session.get(
                config.CAPTCHA_URL,
                headers=headers,
                timeout=config.REQUEST_TIMEOUT
            )
            
            # Check for 401 before raising for status
            if response.status_code == 401:
                logger.error("❌ 401 Unauthorized while fetching CAPTCHA")
                return 'unauthorized', None
            
            response.raise_for_status()
            
            data = response.json()
            
            if data.get('success'):
                captcha_id = data['data']['guid']
                captcha_image = data['data']['captcha']
                logger.debug(f"CAPTCHA fetched: ID={captcha_id}")
                return captcha_id, captcha_image
            else:
                logger.error(f"Failed to get CAPTCHA: {data}")
                return None, None
        except Exception as e:
            logger.error(f"Error fetching CAPTCHA: {e}")
            return None, None
    
    def solve_captcha_with_retry(self, max_retries=None):
        """
        Fetch and solve CAPTCHA with retry logic
        
        Args:
            max_retries (int): Maximum retry attempts
            
        Returns:
            tuple: (captcha_id, captcha_value) or (None, None) on failure
        """
        max_retries = max_retries or 1  # Single attempt only
        
        for attempt in range(max_retries):
            captcha_id, captcha_image = self.get_captcha()
            
            if not captcha_id:
                logger.debug(f"CAPTCHA fetch failed (attempt {attempt + 1}/{max_retries})")
                continue
            
            # Skip debug saves for speed

            # Use Gemini or CaptchaSolver based on configuration
            if self.gemini_solver:
                captcha_value = self.gemini_solver.solve(captcha_image)
            elif self.captcha_solver:
                captcha_value = self.captcha_solver.solve(
                    captcha_image,
                    preprocess=config.CAPTCHA_PREPROCESS
                )
            else:
                logger.error("No OCR solver available")
                return None, None
            
            if captcha_value and len(captcha_value) >= 5:
                logger.info(f"CAPTCHA solved: {captcha_value}")
                return captcha_id, captcha_value
            else:
                logger.warning(f"CAPTCHA solve failed (attempt {attempt + 1}/{max_retries})")
        
        logger.error("Failed to solve CAPTCHA after all retries")
        return None, None
    
    def handle_session_expired(self):
        """
        Handle expired session by:
        1. Starting Chrome with debugging
        2. Waiting for user to login
        3. Extracting fresh cookies
        4. Updating config and session
        
        Returns:
            bool: True if session refreshed, False if failed
        """
        logger.warning("\n" + "=" * 80)
        logger.warning("⚠️  SESSION EXPIRED - Authentication Required")
        logger.warning("=" * 80)
        
        # Increment unauthorized count
        self.unauthorized_count += 1
        
        if self.unauthorized_count > 3:
            logger.error("Too many unauthorized errors. Please check your credentials manually.")
            return False
        
        logger.info("\n🔄 Starting automatic session refresh...")
        logger.info("\nSteps:")
        logger.info("  1. Starting Chrome with debugging mode...")
        logger.info("  2. You will need to login to my3.soliq.uz")
        logger.info("  3. After login, press Enter to continue")
        logger.info("  4. Cookies will be extracted automatically")
        logger.info("  5. Automation will resume")
        logger.info("\n" + "-" * 80)
        
        # Start Chrome with debugging
        try:
            bat_file = Path("START_CHROME_DEBUG.bat")
            if bat_file.exists():
                logger.info("\n▶️  Launching Chrome...")
                subprocess.Popen([str(bat_file)], shell=True)
                time.sleep(5)  # Give Chrome time to start
            else:
                logger.warning("⚠️  START_CHROME_DEBUG.bat not found")
                logger.info("Please start Chrome manually with: --remote-debugging-port=9222")
        except Exception as e:
            logger.error(f"Error starting Chrome: {e}")
        
        # Wait for user to login
        logger.info("\n" + "=" * 80)
        logger.warning("⏸️  AUTOMATION PAUSED")
        logger.info("=" * 80)
        logger.info("\n📌 Please:")
        logger.info("   1. Login to my3.soliq.uz in the Chrome window")
        logger.info("   2. Navigate to the checks page")
        logger.info("   3. Press ENTER here when ready")
        logger.info("\n" + "=" * 80)
        
        try:
            input("\n👉 Press ENTER after you've logged in: ")
        except KeyboardInterrupt:
            logger.info("\n❌ Cancelled by user")
            return False
        
        # Extract cookies
        logger.info("\n🔄 Extracting fresh cookies...")
        
        try:
            # Try to run refresh_cookies.py
            refresh_script = Path("refresh_cookies.py")
            if refresh_script.exists():
                logger.info("Running refresh_cookies.py...")
                result = subprocess.run(
                    [sys.executable, str(refresh_script)],
                    capture_output=True,
                    text=True,
                    timeout=30
                )
                
                if result.returncode == 0:
                    logger.info("✓ Cookies extracted successfully")
                    
                    # Reload config to get new cookies
                    import importlib
                    importlib.reload(config)
                    
                    # Update session cookies
                    if config.SESSION_COOKIES:
                        self.session.cookies.update(config.SESSION_COOKIES)
                        logger.info("✓ Session cookies updated")
                    
                    # Reset unauthorized count
                    self.unauthorized_count = 0
                    self.session_expired = False
                    
                    logger.info("\n" + "=" * 80)
                    logger.info("✅ SESSION REFRESHED - Resuming automation")
                    logger.info("=" * 80 + "\n")
                    
                    time.sleep(2)
                    return True
                else:
                    logger.error(f"Cookie extraction failed: {result.stderr}")
            else:
                logger.warning("refresh_cookies.py not found")
                logger.info("\nPlease update config.py manually with fresh cookies.")
                input("Press ENTER when config.py is updated: ")
                
                # Reload config
                import importlib
                importlib.reload(config)
                
                if config.SESSION_COOKIES:
                    self.session.cookies.update(config.SESSION_COOKIES)
                    self.unauthorized_count = 0
                    self.session_expired = False
                    logger.info("✓ Session updated manually")
                    return True
        
        except Exception as e:
            logger.error(f"Error during session refresh: {e}")
        
        return False
    
    def submit_check_edit(self, check_data, captcha_id, captcha_value, file_id=""):
        """
        Submit check edit to server with retry logic for error 9999
        
        Requires Bearer token authentication (same as file upload).
        
        Args:
            check_data (dict): Check data from Excel
            captcha_id (str): CAPTCHA ID
            captcha_value (str): Solved CAPTCHA value
            file_id (str): Uploaded file ID
            
        Returns:
            bool or str: True if successful, False on failure, 'captcha_error' for CAPTCHA errors
        """
        import random
        
        # Retry configuration for server error 9999
        max_retries_9999 = getattr(config, 'SERVER_ERROR_MAX_RETRIES', 5)
        base_delay_9999 = getattr(config, 'SERVER_ERROR_BASE_DELAY', 2.0)
        max_delay_9999 = getattr(config, 'SERVER_ERROR_MAX_DELAY', 30.0)
        
        for retry_attempt in range(max_retries_9999):
            try:
                # Prepare payload matching HAR file structure exactly
                payment_id = check_data['payment_id']
                payment_date = check_data.get('payment_date', '2025-10-01')
                tin = check_data.get('tin', config.DEFAULT_TIN)
                terminal_id = check_data.get('terminal_id', 'EP000000000551')
                
                payload = {
                    "paymentId": payment_id,
                    "vatTotal": 0,  # Set to zero as per requirements
                    "cashTotal": 0,  # Set to zero as per requirements
                    "cardTotal": "0",  # Set to zero as per requirements
                    "attachedFile": file_id if file_id else "",  # Handle empty file ID
                    "captchaId": captcha_id,
                    "captchaValue": captcha_value,
                    "cardType": "",  # Empty string as per HAR
                    "nameStatus": True,  # Boolean as per HAR
                    "clientIp": "77.77.777.7",  # IP from HAR file
                    "paymentDetails": [  # FIXED: Complete structure matching successful HAR
                        {
                            "id": f"{payment_id}-0",  # Required: payment_id + "-0"
                            "paymentId": payment_id,  # Required: same as parent
                            "tin": tin,  # Required: TIN
                            "pinfl": None,  # Required: null
                            "name": f"{check_data['check_number']}-check edit",  # Required: description
                            "price": 0,
                            "vat": "0",
                            "amount": 0,
                            "discount": 0,
                            "other": 0,
                            "barCode": None,
                            "label": None,
                            "productCode": "10701001018000000",
                            "unitCode": None,
                            "unitName": None,
                            "vatPercent": "0",
                            "commissionTin": config.DEFAULT_TIN,
                            "commissionPinfl": None,
                            "packageCode": "1495029",
                            "paymentDate": payment_date,  # Required: payment date inside details
                            "terminalId": None,  # Required: null in HAR
                            "year": None,
                            "month": None,
                            "day": None,
                            "terminalStateId": None,
                            "isRefund": None,
                            "existsCommission": None,
                            "vaucher": 0,
                            "isNotLabel": None
                        }
                    ],
                    "tin": tin,
                    "terminalId": terminal_id,
                    "paymentDate": payment_date
                }
                
                logger.debug(f"Submitting check edit: {check_data['check_number']}")
                
                # Prepare headers with Bearer token
                headers = dict(self.session.headers)
                bearer_token = self._get_bearer_token()

                if bearer_token:
                    headers['Authorization'] = bearer_token
                    logger.debug(f"Using Bearer token for check submission")
                else:
                    logger.warning("⚠️ No Bearer token - check submission may fail!")
                
                # Serialize payload manually
                payload_str = json.dumps(payload)
                
                response = self.session.post(
                    config.SUBMIT_URL,
                    data=payload_str,
                    headers=headers,
                    timeout=config.REQUEST_TIMEOUT
                )
                
                # Check for 401 Unauthorized
                if response.status_code == 401:
                    logger.error("❌ 401 Unauthorized - Session expired")
                    return 'unauthorized'
                
                response.raise_for_status()
                
                result = response.json()
                
                if result.get('success'):
                    logger.info(f"[OK] Check {check_data['check_number']} submitted successfully")
                    return True
                else:
                    reason = result.get('reason', result.get('message', 'Unknown error'))
                    error_code = result.get('code') or (reason.get('code') if isinstance(reason, dict) else None)
                    
                    # Check if it's a CAPTCHA error
                    if error_code == '1018' or 'captcha' in str(reason).lower():
                        safe_reason = json.dumps(reason, ensure_ascii=True)
                        logger.warning(f"CAPTCHA error for check {check_data['check_number']}: {safe_reason}")
                        
                        # Save failed CAPTCHA for debugging
                        try:
                            if Path("last_captcha.png").exists():
                                debug_filename = f"failed_captcha_{check_data['check_number']}_{captcha_value}_{int(time.time())}.png"
                                shutil.copy("last_captcha.png", debug_filename)
                                logger.info(f"Saved failed CAPTCHA to {debug_filename}")
                        except Exception as e:
                            logger.debug(f"Failed to save debug captcha: {e}")
                            
                        return 'captcha_error'
                    
                    # Check if it's a duplicate error (Code 9099)
                    # "Ushbu check uchun hali ko'rib chiqilmagan tahrirlash arizasi mavjud!"
                    # This means we already submitted it, so we should treat it as success
                    elif error_code == '9099':
                        logger.info(f"[OK] Check {check_data['check_number']} already submitted (Duplicate 9099) - Marking as done")
                        return True
                    
                    # Check if it's SERVER ERROR 9999 - retry with exponential backoff
                    elif error_code == '9999':
                        if retry_attempt < max_retries_9999 - 1:
                            # Calculate exponential backoff delay with jitter
                            delay = min(base_delay_9999 * (2 ** retry_attempt), max_delay_9999)
                            jitter = random.uniform(0.5, 1.5)
                            actual_delay = delay * jitter
                            
                            logger.warning(f"⚠️ Server error 9999 for check {check_data['check_number']} - "
                                         f"Retry {retry_attempt + 1}/{max_retries_9999} in {actual_delay:.1f}s")
                            time.sleep(actual_delay)
                            continue  # Retry the submission
                        else:
                            logger.error(f"❌ Server error 9999 persisted after {max_retries_9999} retries for check {check_data['check_number']}")
                            return 'server_error_9999'
                        
                    else:
                        logger.error(f"Failed to submit check {check_data['check_number']}: {reason}")
                        logger.error(f"Full response: {result}")
                        logger.error(f"Payload sent: {json.dumps(payload)}")
                        logger.error(f"Headers used: {headers}")
                        return False
                        
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
                if retry_attempt < max_retries_9999 - 1:
                    delay = min(base_delay_9999 * (2 ** retry_attempt), max_delay_9999)
                    jitter = random.uniform(0.5, 1.5)
                    actual_delay = delay * jitter
                    logger.warning(f"⚠️ Connection error for check {check_data['check_number']}: {e} - "
                                 f"Retry {retry_attempt + 1}/{max_retries_9999} in {actual_delay:.1f}s")
                    time.sleep(actual_delay)
                    continue
                else:
                    logger.error(f"Connection error persisted after {max_retries_9999} retries: {e}")
                    return False
            except Exception as e:
                logger.error(f"Error submitting check: {e}")
                return False
        
        # All retries exhausted
        return False
    
    def process_batch(self, checks_chunk):
        """
        Process a batch of checks using Gemini Batch OCR with PARALLEL execution
        """
        logger.info(f"Processing batch of {len(checks_chunk)} checks (Parallel)...")
        
        # 1. Fetch CAPTCHAs and Upload Files in PARALLEL
        batch_data = []
        images_b64 = []
        skipped_count = 0
        failed_checks = []
        
        # Helper function for parallel execution with fixed delay (smoother CPU usage)
        upload_delay = getattr(config, 'FILE_UPLOAD_STAGGER_DELAY', 0.2)
        
        def prepare_check(index, check_data):
            check_number = check_data['check_number']
            
            # Simple fixed delay for each request (smoother than position-based)
            if upload_delay > 0:
                time.sleep(upload_delay)
            
            # Skip if already processed (double check)
            if check_number in self.progress:
                return None
            
            # Upload file
            file_id = self.upload_file(check_number)
            if file_id == 'unauthorized':
                return 'unauthorized'
            
            # Fetch CAPTCHA
            captcha_id, captcha_image = self.get_captcha()
            if captcha_id == 'unauthorized':
                return 'unauthorized'
                
            if captcha_id and captcha_image:
                return {
                    'check_data': check_data,
                    'file_id': file_id,
                    'captcha_id': captcha_id,
                    'captcha_image': captcha_image
                }
            return False

        # Execute preparation in parallel
        prepared_items = []
        max_workers_upload = config.MAX_WORKERS_FILE_UPLOAD
        logger.debug(f"Using {max_workers_upload} parallel workers for file uploads")
        
        with ThreadPoolExecutor(max_workers=max_workers_upload) as executor:
            future_to_check = {executor.submit(prepare_check, i, c): c for i, c in enumerate(checks_chunk)}
            
            for future in as_completed(future_to_check):
                check = future_to_check[future]
                try:
                    result = future.result()
                    if result == 'unauthorized':
                        logger.error(f"401 Unauthorized during prep for {check['check_number']}")
                        self.session_expired = True
                        return [] # Stop batch
                    elif result:
                        prepared_items.append(result)
                    elif result is None:
                        skipped_count += 1
                    else:
                        logger.error(f"Failed prep for {check['check_number']}")
                        failed_checks.append(check)
                except Exception as e:
                    logger.error(f"Exception during prep for {check['check_number']}: {e}")
                    failed_checks.append(check)

        # Sort to maintain some order (optional but good for logs)
        prepared_items.sort(key=lambda x: x['check_data']['row_index'])
        
        # Extract images for Gemini
        for item in prepared_items:
            batch_data.append(item)
            images_b64.append(item['captcha_image'])
            
        if skipped_count > 0:
            logger.info(f"Skipped {skipped_count} checks")
            
        if not batch_data:
            logger.info("No checks to process in this batch")
            return failed_checks
            
        # 2. Solve batch with Gemini (Single call, fast)
        if self.gemini_solver:
            logger.info(f"Solving {len(images_b64)} CAPTCHAs with Gemini...")
            solved_captchas = self.gemini_solver.solve_batch(images_b64)
        else:
            logger.error("Gemini solver not initialized")
            return failed_checks
        
        # 3. Submit checks in PARALLEL
        logger.info(f"Submitting {len(batch_data)} checks in parallel...")
        
        max_workers_submit = config.MAX_WORKERS_CHECK_SUBMIT
        logger.debug(f"Using {max_workers_submit} parallel workers for check submissions")
        
        
        # Fixed delay between submissions (smoother CPU usage)
        submission_delay = getattr(config, 'SUBMISSION_STAGGER_DELAY', 0.15)
        
        def submit_check(i, item):
            check_data = item['check_data']
            
            # Simple fixed delay for each submission (smoother than position-based)
            if submission_delay > 0:
                time.sleep(submission_delay)
            
            if i >= len(solved_captchas):
                return False
                
            captcha_value = solved_captchas[i]
            if not captcha_value:
                return 'captcha_fail'
                
            return self.submit_check_edit(
                check_data, 
                item['captcha_id'], 
                captcha_value, 
                item['file_id']
            )

        new_completed_checks = []
        
        with ThreadPoolExecutor(max_workers=max_workers_submit) as executor:
            # Map futures to items
            future_to_item = {executor.submit(submit_check, i, item): item for i, item in enumerate(batch_data)}
            
            for future in as_completed(future_to_item):
                item = future_to_item[future]
                check_data = item['check_data']
                try:
                    result = future.result()
                    
                    if result is True:
                        self.progress.add(check_data['check_number'])
                        new_completed_checks.append(check_data['check_number'])
                        self.stats['successful'] += 1
                        self.delete_failed_captcha_images(check_data['check_number'])
                        # logger.info(f"✓ {check_data['check_number']}") # Reduce log spam
                    elif result == 'unauthorized':
                        self.session_expired = True
                        # Queue remaining? The loop continues but next batches will fail
                    elif result == 'captcha_fail' or result == 'captcha_error':
                        self.stats['captcha_failures'] += 1
                        failed_checks.append(check_data)
                    else:
                        self.stats['failed'] += 1
                        failed_checks.append(check_data)
                        
                except Exception as e:
                    logger.error(f"Error submitting {check_data['check_number']}: {e}")
                    failed_checks.append(check_data)
        
        # Save progress (Incremental)
        if new_completed_checks:
            self.save_progress(new_checks=new_completed_checks)
            
        return failed_checks

    def process_check(self, check_data):
        """
        Process a single check with complete workflow
        
        Args:
            check_data (dict): Check data from Excel
            
        Returns:
            bool: True if successful, False otherwise
        """
        # Add random delay to prevent workers from synchronizing
        import random
        delay_config = getattr(config, 'DELAY_BETWEEN_CHECKS', 0.5)
        delay = random.uniform(delay_config * 0.8, delay_config * 1.5)
        time.sleep(delay)
        
        check_number = check_data['check_number']
        
        # Skip if already processed
        if check_number in self.progress:
            logger.debug(f"Skipping already processed check: {check_number}")
            return True
        
        # Check if session is expired and handle it
        if self.session_expired:
            if not self.handle_session_expired():
                logger.error("Failed to refresh session. Stopping.")
                return False
        
        # Upload file
        file_id = self.upload_file(check_number)
        if not file_id and self.zip_file_content:
             logger.warning(f"Proceeding with empty file ID for check {check_number} (upload failed or skipped)")

        
        # Process with retries
        for attempt in range(config.MAX_RETRIES):
            # Solve CAPTCHA
            captcha_id, captcha_value = self.solve_captcha_with_retry(max_retries=config.CAPTCHA_MAX_RETRIES)
            
            if not captcha_id:
                logger.error(f"Failed to solve CAPTCHA for check {check_number}")
                continue
            
            # Submit check edit
            result = self.submit_check_edit(check_data, captcha_id, captcha_value, file_id)
            
            # Check if submission returned unauthorized
            if result == 'unauthorized':
                logger.warning("Submission failed due to auth - triggering session refresh")
                self.session_expired = True
                return False  # Will retry on next attempt
            
            if result is True:
                # Success
                self.progress.add(check_number)
                self.stats['successful'] += 1
                
                # Delete any failed CAPTCHA images for this check (in case it was a retry)
                self.delete_failed_captcha_images(check_number)
                
                # Save progress periodically
                if len(self.progress) % config.SAVE_PROGRESS_EVERY == 0:
                    self.save_progress()
                
                return True
            elif result == 'captcha_error':
                # CAPTCHA error - retry with new CAPTCHA
                self.stats['captcha_failures'] += 1
                logger.warning(f"CAPTCHA error, retrying check {check_number} (attempt {attempt + 1}/{config.MAX_RETRIES})")
                continue
            else:
                # Other error - retry
                logger.warning(f"Submit failed, retrying check {check_number} (attempt {attempt + 1}/{config.MAX_RETRIES})")
                continue
        
        # All retries failed
        logger.error(f"[X] Failed to process check {check_number} after {config.MAX_RETRIES} attempts")
        self.stats['failed'] += 1
        return False
    
    def load_checks_from_excel(self):
        """
        Load check data from Excel file with enhanced data extraction
        
        Returns:
            list: List of check data dictionaries
        """
        try:
            logger.info(f"Loading Excel file: {config.EXCEL_FILE}")
            wb = load_workbook(config.EXCEL_FILE, read_only=True)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1]]
            logger.info(f"Excel columns: {headers}")
            
            # Find check number column
            check_col_idx = None
            if isinstance(config.CHECK_NUMBER_COLUMN, int):
                check_col_idx = config.CHECK_NUMBER_COLUMN
            else:
                for idx, header in enumerate(headers):
                    if header and config.CHECK_NUMBER_COLUMN in str(header):
                        check_col_idx = idx
                        break
            
            if check_col_idx is None:
                logger.error(f"Could not find check number column: {config.CHECK_NUMBER_COLUMN}")
                return []
            
            logger.info(f"Check number column index: {check_col_idx}")
            
            # Find other useful columns
            payment_id_col = check_col_idx  # Assuming check number is part of payment ID
            
            # Load checks
            checks = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # --- FIX START ---
                # 1. Extract the Check Number first
                raw_check_num = row[check_col_idx]
                if not raw_check_num:
                    continue  # Skip empty rows
                
                check_number = str(raw_check_num).strip()
                
                # 2. Generate Payment ID with correct format from HAR analysis
                # Format: terminalId + (padded check number to 16 digits)
                # Total length must be 30 chars. Terminal is 14 chars. Check part must be 16 chars.
                try:
                    terminal_id_prefix = row[headers.index('terminalId')] if 'terminalId' in headers else 'EP000000000551'
                except ValueError:
                    terminal_id_prefix = 'EP000000000551'  # Default from HAR
                
                # Pad check number with zeros on the left to make it 16 digits long
                # Example 11-digit: 26371353560 -> 0000026371353560 (5 zeros)
                # Example 10-digit: 5464588689 -> 0000005464588689 (6 zeros)
                check_number_part = check_number.zfill(16)
                payment_id = terminal_id_prefix + check_number_part
                # --- FIX END ---

                try:
                    product_code = row[headers.index('productCode')] if 'productCode' in headers else "10701001018000000"
                    package_code = row[headers.index('packageCode')] if 'packageCode' in headers else "1495029"
                    commission_tin = row[headers.index('commissionTin')] if 'commissionTin' in headers else "62409036610049"
                except (ValueError, IndexError):
                    # Use defaults if columns not found
                    product_code = "10701001018000000"
                    package_code = "1495029"
                    commission_tin = "62409036610049"
                
                payment_details = [{
                    "price": 0,
                    "amount": 0,
                    "vatPercent": "0",
                    "vat": "0",
                    "discount": 0,
                    "other": 0,
                    "productCode": str(product_code),
                    "vaucher": 0,
                    "packageCode": str(package_code),
                    "unitName": None,
                    "commissionTin": str(commission_tin),
                    "commissionPinfl": ""
                }]

                # Extract payment date if available
                payment_date = "2025-10-12"  # Default from HAR
                try:
                    date_idx = headers.index('paymentDate') if 'paymentDate' in headers else (headers.index('date') if 'date' in headers else -1)
                    if date_idx != -1 and row[date_idx]:
                        raw_date = str(row[date_idx])
                        # Format: 2025-10-11 20:48:55.126844 -> 2025-10-11
                        if ' ' in raw_date:
                            payment_date = raw_date.split(' ')[0]
                        else:
                            payment_date = raw_date[:10]
                except ValueError:
                    # Date column not found, use default
                    pass

                # Try to extract tin from Excel if available
                try:
                    tin = row[headers.index('tin')] if 'tin' in headers else '304739262'
                except ValueError:
                    tin = '304739262'  # Default from HAR
                
                try:
                    terminal_id = row[headers.index('terminalId')] if 'terminalId' in headers else 'EP000000000551'
                except ValueError:
                    terminal_id = 'EP000000000551'  # Default from HAR
                
                check_data = {
                    'check_number': check_number,
                    'payment_id': payment_id,
                    'row_index': row_idx,
                    'payment_details': payment_details,
                    'tin': str(tin),  # Ensure it's a string
                    'terminal_id': str(terminal_id),  # Ensure it's a string
                    'payment_date': payment_date
                }
                
                checks.append(check_data)
            wb.close()
            logger.info(f"Loaded {len(checks)} valid checks from Excel")
            return checks
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            return []
    
    def mark_completed_in_excel(self, check_data):
        """
        Mark check as completed in Excel file
        
        Args:
            check_data (dict): Check data with row_index
        """
        try:
            wb = load_workbook(config.EXCEL_FILE)
            ws = wb.active
            
            # Find or create status column
            headers = [cell.value for cell in ws[1]]
            status_col_idx = None
            
            for idx, header in enumerate(headers):
                if header and config.STATUS_COLUMN in str(header):
                    status_col_idx = idx + 1
                    break
            
            if status_col_idx is None:
                # Create status column
                status_col_idx = len(headers) + 1
                ws.cell(row=1, column=status_col_idx, value=config.STATUS_COLUMN)
                logger.info(f"Created status column at index {status_col_idx}")
            
            # Mark as completed
            ws.cell(row=check_data['row_index'], column=status_col_idx, value=config.COMPLETED_STATUS)
            
            wb.save(config.EXCEL_FILE)
            wb.close()
            
            logger.debug(f"Marked check {check_data['check_number']} as completed in Excel")
        except Exception as e:
            logger.error(f"Error marking check in Excel: {e}")
    
    def run(self, limit=None, dry_run=False):
        """
        Run the automation
        
        Args:
            limit (int): Maximum number of checks to process (for testing)
            dry_run (bool): If True, don't actually submit changes
        """
        logger.info("=" * 80)
        logger.info("Check Automation Script Started - Enhanced Version")
        logger.info(f"Dry run: {dry_run}")
        logger.info(f"Limit: {limit if limit else 'None (all checks)'}")
        logger.info("=" * 80)
        
        self.stats['start_time'] = datetime.now()
        
        # Start cookie refresher
        if not dry_run:
            self.cookie_refresher.start()
        
        try:
            # Load checks
            checks = self.load_checks_from_excel()
            
            if not checks:
                logger.error("No checks loaded. Exiting.")
                return
            
            # Apply limit if specified
            if limit:
                checks = checks[:limit]
                logger.info(f"Processing limited to {limit} checks")
            
            # Process checks
            # Process checks
            if self.gemini_solver:
                # Filter out already processed checks to ensure full batches
                pending_checks = [c for c in checks if c['check_number'] not in self.progress]
                skipped_count = len(checks) - len(pending_checks)
                
                if skipped_count > 0:
                    logger.info(f"Skipping {skipped_count} already completed checks")
                
                if not pending_checks:
                    logger.info("All checks already completed!")
                    return

                # Initialize retry counts
                for c in pending_checks:
                    c['retry_count'] = 0

                # Batch processing with queue
                batch_size = config.BATCH_SIZE
                batch_num = 0
                
                while pending_checks:
                    batch_num += 1
                    # Take chunk
                    chunk = pending_checks[:batch_size]
                    pending_checks = pending_checks[batch_size:]
                    
                    # Estimate total batches (dynamic)
                    total_remaining = len(pending_checks) + len(chunk)
                    est_batches = (total_remaining + batch_size - 1) // batch_size + (batch_num - 1)

                    logger.info(f"\n[{batch_num}/~{est_batches}] Processing batch of {len(chunk)} checks")
                    
                    if dry_run:
                        logger.info("DRY RUN: Skipping batch")
                        time.sleep(0.5)
                        continue
                        
                    failed_captcha_checks = self.process_batch(chunk)
                    
                    # Check if session expired during batch processing
                    if self.session_expired:
                        logger.warning("\n" + "="*80)
                        logger.warning("⚠️  Session expired detected during batch processing")
                        logger.warning("="*80)
                        
                        # Try to refresh session
                        if self.handle_session_expired():
                            # Re-queue all failed checks including current chunk
                            logger.info("Session refreshed successfully, re-queueing current batch")
                            pending_checks = failed_captcha_checks + pending_checks
                            continue
                        else:
                            logger.error("Failed to refresh session. Stopping automation.")
                            return
                    
                    # Handle retries
                    requeue_checks = []
                    for failed_check in failed_captcha_checks:
                        failed_check['retry_count'] += 1
                        if failed_check['retry_count'] <= config.CAPTCHA_MAX_RETRIES:
                            requeue_checks.append(failed_check)
                        else:
                            logger.error(f"Check {failed_check['check_number']} failed CAPTCHA after {failed_check['retry_count']} attempts")
                            self.stats['failed'] += 1
                    
                    if requeue_checks:
                        logger.info(f"Re-queueing {len(requeue_checks)} checks for next batch")
                        pending_checks = requeue_checks + pending_checks
                    
                    self.stats['processed'] += len(chunk) - len(failed_captcha_checks) # Only count processed (success or hard fail)
                    # Note: We don't count requeued checks as "processed" yet because they are going back to queue
                    
                    # Progress report
                    elapsed = (datetime.now() - self.stats['start_time']).total_seconds()
                    rate = self.stats['successful'] / elapsed if elapsed > 0 else 0 # Use successful for rate
                    
                    # Calculate total progress
                    total_completed = len(self.progress)
                    
                    logger.info(f"Progress: {total_completed}/{len(checks)} ({total_completed/len(checks)*100:.1f}%)")
                    logger.info(f"Success: {self.stats['successful']}, Failed: {self.stats['failed']}, Re-queued: {len(requeue_checks)}")
                    logger.info(f"Rate: {rate:.2f} success/sec")
            else:
                # Sequential processing
                for idx, check_data in enumerate(checks, start=1):
                    logger.info(f"\n[{idx}/{len(checks)}] Processing check: {check_data['check_number']}")
                    
                    if dry_run:
                        logger.info("DRY RUN: Skipping actual submission")
                        time.sleep(0.5)
                        continue
                    
                    # Process the check
                    success = self.process_check(check_data)
                    
                    if success:
                        # Mark in Excel
                        self.mark_completed_in_excel(check_data)
                    
                    self.stats['processed'] += 1
                    
                    # Progress report
                    if idx % 10 == 0:
                        elapsed = (datetime.now() - self.stats['start_time']).total_seconds()
                        rate = self.stats['processed'] / elapsed if elapsed > 0 else 0
                        remaining = len(checks) - idx
                        eta_seconds = remaining / rate if rate > 0 else 0
                        eta_hours = eta_seconds / 3600
                        
                        logger.info(f"\nProgress: {idx}/{len(checks)} ({idx/len(checks)*100:.1f}%)")
                        logger.info(f"Success: {self.stats['successful']}, Failed: {self.stats['failed']}")
                        logger.info(f"Rate: {rate:.2f} checks/sec, ETA: {eta_hours:.1f} hours")
                    
                    # Delay between requests
                    time.sleep(config.REQUEST_DELAY)
        
        finally:
            # Stop cookie refresher
            self.cookie_refresher.stop()
            
            # Final statistics
            self.stats['end_time'] = datetime.now()
            
            # Cleanup CAPTCHA images
            try:
                # Delete last_captcha.png
                if Path("last_captcha.png").exists():
                    Path("last_captcha.png").unlink()
                
                # Delete failed CAPTCHA images
                for p in Path(".").glob("failed_captcha_*.png"):
                    p.unlink()
                
                logger.info("Cleaned up CAPTCHA images")
            except Exception as e:
                logger.warning(f"Error cleaning up CAPTCHA images: {e}")
            self.save_progress()
            self.print_statistics()
    
    def print_statistics(self):
        """Print final statistics"""
        logger.info("\n" + "=" * 80)
        logger.info("AUTOMATION COMPLETED")
        logger.info("=" * 80)
        logger.info(f"Total checks: {self.stats['total']}")
        logger.info(f"Processed: {self.stats['processed']}")
        logger.info(f"Successful: {self.stats['successful']}")
        logger.info(f"Failed: {self.stats['failed']}")
        logger.info(f"CAPTCHA failures: {self.stats['captcha_failures']}")
        
        if self.stats['start_time'] and self.stats['end_time']:
            duration = self.stats['end_time'] - self.stats['start_time']
            logger.info(f"Duration: {duration}")
            
            if self.stats['processed'] > 0:
                avg_time = duration.total_seconds() / self.stats['processed']
                logger.info(f"Average time per check: {avg_time:.2f} seconds")
        
        success_rate = (self.stats['successful'] / self.stats['processed'] * 100) if self.stats['processed'] > 0 else 0
        logger.info(f"Success rate: {success_rate:.1f}%")
        logger.info("=" * 80)


def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Automate check editing on my3.soliq.uz')
    parser.add_argument('--limit', type=int, help='Limit number of checks to process (for testing)')
    parser.add_argument('--dry-run', action='store_true', help='Dry run mode (no actual submissions)')
    
    args = parser.parse_args()
    
    try:
        automation = CheckAutomation()
        automation.run(limit=args.limit, dry_run=args.dry_run)
    except KeyboardInterrupt:
        logger.info("\nInterrupted by user. Progress has been saved.")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
