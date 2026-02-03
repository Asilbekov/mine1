"""
Redistribute progress entries to correct files.
Reads all progress files and Excel files to determine which checks belong where.
"""

import json
import sqlite3
from pathlib import Path
from openpyxl import load_workbook
import sys

# Redirect output to file to avoid encoding issues
log_file = open("redistribute_log.txt", "w", encoding="utf-8")

def log(msg):
    log_file.write(msg + "\n")
    log_file.flush()

# Excel files and their corresponding progress files
file_mappings = [
    ("Асилбекова_Сентябрь_тахрирлаш_42930.xlsx", "progress_Асилбекова_Сентябрь_тахрирлаш_42930"),
    ("Асилбекова_Август_тахрирлаш_48606_та.xlsx", "progress_Асилбекова_Август_тахрирлаш_48606_та"),
    ("Асилбекова Июнь тахрирлаш 20303 та.xlsx", "progress_Асилбекова Июнь тахрирлаш 20303 та"),
    ("Асилбекова Июль тахрирлаш 46430 та.xlsx", "progress_Асилбекова Июль тахрирлаш 46430 та"),
]

log("Step 1: Loading all check numbers from Excel files...")
log("")

# Build mapping: check_number -> excel_file
check_to_file = {}

for excel_file, _ in file_mappings:
    if not Path(excel_file).exists():
        log(f"Warning: {excel_file} not found, skipping")
        continue
        
    wb = load_workbook(excel_file, read_only=True)
    ws = wb.active
    
    # Find receipt_id column
    headers = [cell.value for cell in ws[1]]
    check_col_idx = None
    for idx, header in enumerate(headers):
        if header and 'receipt_id' in str(header):
            check_col_idx = idx
            break
    
    if check_col_idx is None:
        log(f"Warning: receipt_id column not found in {excel_file}")
        continue
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        check_num = row[check_col_idx]
        if check_num:
            check_to_file[str(check_num).strip()] = excel_file
            count += 1
    
    wb.close()
    log(f"  {excel_file}: {count} checks")

log(f"")
log(f"Total unique checks loaded: {len(check_to_file)}")
log("")

log("Step 2: Collecting all completed checks from progress files...")
log("")

# Collect all completed checks from all progress files
all_completed = set()

for _, progress_base in file_mappings:
    # Check JSON
    json_file = f"{progress_base}.json"
    if Path(json_file).exists():
        with open(json_file, 'r', encoding='utf-8') as f:
            try:
                checks = json.load(f)
                all_completed.update(checks)
                log(f"  {json_file}: {len(checks)} checks")
            except:
                log(f"  {json_file}: corrupted, skipping")
    
    # Check DB
    db_file = f"{progress_base}.db"
    if Path(db_file).exists():
        conn = sqlite3.connect(db_file)
        cursor = conn.execute("SELECT check_number FROM progress")
        db_checks = [row[0] for row in cursor.fetchall()]
        all_completed.update(db_checks)
        conn.close()
        log(f"  {db_file}: {len(db_checks)} checks")

log("")
log(f"Total completed checks found: {len(all_completed)}")
log("")

log("Step 3: Redistributing checks to correct files...")
log("")

# Group completed checks by their correct file
file_to_checks = {excel: set() for excel, _ in file_mappings}

orphaned = []
for check in all_completed:
    if check in check_to_file:
        target_file = check_to_file[check]
        file_to_checks[target_file].add(check)
    else:
        orphaned.append(check)

if orphaned:
    log(f"Warning: {len(orphaned)} checks not found in any Excel file (might be from old data)")

# Write to correct files
for excel_file, progress_base in file_mappings:
    checks = sorted(file_to_checks[excel_file])
    
    if not checks:
        log(f"  {excel_file}: 0 checks (skipping)")
        continue
    
    # Backup old files
    json_file = f"{progress_base}.json"
    db_file = f"{progress_base}.db"
    
    if Path(json_file).exists():
        Path(json_file).rename(f"{json_file}.backup")
    if Path(db_file).exists():
        Path(db_file).rename(f"{db_file}.backup")
    
    # Write JSON
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(checks, f, ensure_ascii=False, indent=2)
    
    # Write DB
    conn = sqlite3.connect(db_file)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS progress (
            check_number TEXT PRIMARY KEY,
            completed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.executemany(
        "INSERT OR IGNORE INTO progress (check_number) VALUES (?)",
        [(c,) for c in checks]
    )
    conn.commit()
    conn.close()
    
    log(f"  {excel_file}: {len(checks)} checks written")

log("")
log("Done! Old files backed up with .backup extension")
log("You can now restart master_controller.py")

log_file.close()
print("Redistribution complete! Check redistribute_log.txt for details")

