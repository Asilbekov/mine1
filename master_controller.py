"""
Master Controller for Multi-Worker Automation
Supports distributing workers across multiple Excel files.

This script:
1. Refreshes cookies ONCE
2. Starts 8 workers (2 per file initially)
3. Monitors progress of all files
4. Reassigns workers from completed files to remaining ones
"""

import subprocess
import threading
import time
import json
import sys
import argparse
from pathlib import Path
from datetime import datetime
import logging
from progress_db import ProgressTracker
import config

# Hide months that are completed 90%+ from progress display
HIDE_PROGRESS_THRESHOLD = 90  # Percentage

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)


class FileProgressTracker:
    """Tracks progress for a single Excel file"""
    
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.basename = Path(excel_file).stem
        self.db_file = f"progress_{self.basename}.db"
        
        # Calculate total checks
        self.total_checks = self._count_checks()
        logger.info(f"📄 File {self.basename}: {self.total_checks} checks")
        
    def _count_checks(self):
        """Count total checks in Excel file"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_file, read_only=True)
            ws = wb.active
            
            # Find check number column
            headers = [cell.value for cell in ws[1]]
            check_col_idx = None
            if isinstance(config.CHECK_NUMBER_COLUMN, int):
                check_col_idx = config.CHECK_NUMBER_COLUMN
            else:
                for idx, header in enumerate(headers):
                    if header and config.CHECK_NUMBER_COLUMN in str(header):
                        check_col_idx = idx
                        break
            
            if check_col_idx is not None:
                count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[check_col_idx]:
                        count += 1
                wb.close()
                return count
            
            wb.close()
            return ws.max_row - 1
        except Exception as e:
            logger.warning(f"Could not read {self.excel_file}: {e}")
            return 10000  # Fallback
            
    def get_completed_count(self):
        """Get number of completed checks"""
        try:
            # Read directly from SQLite DB for real-time updates
            from progress_db import ProgressTracker
            tracker = ProgressTracker(self.db_file)
            return tracker.count()
        except Exception:
            # Fallback to JSON if DB fails (though unlikely)
            try:
                if Path(self.progress_file).exists():
                    with open(self.progress_file, 'r') as f:
                        data = json.load(f)
                        return len(data)
            except:
                pass
        return 0

    def is_complete(self):
        return self.get_completed_count() >= self.total_checks


class WorkerManager:
    """Manages workers across multiple files"""
    
    def __init__(self, total_workers=8):
        self.total_workers = total_workers
        self.workers = []  # List of dicts: {id, process, file, assigned_id}
        self.files = config.EXCEL_FILES
        self.file_trackers = {f: FileProgressTracker(f) for f in self.files}
        self.running = True
        self.checks_file_set = None
        self.processed_files = set()
        
        # Session activity tracking
        self.session_success_count = 0
        try:
            self.log_handler = open('automation.log', 'r', encoding='utf-8', errors='replace')
            self.log_handler.seek(0, 2) # Start at end
        except:
            self.log_handler = None
        
    def start_initial_distribution(self):
        """Assign workers evenly across files"""
        logger.info("=" * 80)
        logger.info(f"Starting {self.total_workers} workers across {len(self.files)} files...")
        logger.info("=" * 80)
        
        # Sequential Mode: Assign ALL workers to the first file
        if not self.files:
            logger.error("No files to process!")
            return

        first_file = self.files[0]
        logger.info(f"Sequential Mode: Assigning ALL {self.total_workers} workers to {Path(first_file).stem}")
        
        for i in range(self.total_workers):
            worker_id = i + 1  # Fixed: Define worker_id properly
            self.start_worker(worker_id, first_file, worker_id, self.total_workers)
            time.sleep(0.3)
                
    def start_worker(self, global_id, excel_file, local_id, total_local_workers):
        """Start a single worker process"""
        logger.info(f"   Starting Worker {global_id} -> {Path(excel_file).stem} (Local {local_id}/{total_local_workers})")
        
        cmd = [
            sys.executable,
            "automate_worker.py",
            "--worker-id", str(global_id),  # ✅ FIXED: Use global_id for unique API key assignment
            "--total-workers", str(self.total_workers),  # ✅ FIXED: Use total workers, not local count
            "--excel-file", excel_file
        ]
        
        # Pass checks-file if configured
        if hasattr(self, 'checks_file') and self.checks_file:
            cmd.extend(["--checks-file", self.checks_file])
        
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
        )
        
        self.workers.append({
            'global_id': global_id,
            'process': process,
            'file': excel_file,
            'local_id': local_id,
            'total_local_workers': total_local_workers
        })
        
    def stop_worker(self, worker):
        """Stop a specific worker"""
        if worker['process'].poll() is None:
            worker['process'].terminate()
            try:
                worker['process'].wait(timeout=5)  # Increased timeout for graceful shutdown
            except subprocess.TimeoutExpired:
                logger.warning(f"Worker {worker['global_id']} did not terminate gracefully, forcing kill")
                worker['process'].kill()
                worker['process'].wait()  # Ensure process is fully cleaned up
                
    def redistribute_workers(self):
        """Check for completed files and redistribute workers"""
        # Identify completed files
        completed_files = []
        active_files = []
        
        for f in self.files:
            if f in self.processed_files:
                completed_files.append(f)
                continue
                
            tracker = self.file_trackers[f]
            completed = tracker.get_completed_count()
            total = tracker.total_checks
            
            if completed >= total:
                completed_files.append(f)
            else:
                # In Retry Mode, if workers finished (exit code 0) but total < 100%, 
                # it might mean they processed all assigned checks.
                # However, we don't easily track per-file worker history here.
                # But if NO workers are running on this file, and we are here, 
                # we should probably mark it as skipped/done if we are in sequential scan.
                active_files.append(f)

        if hasattr(self, 'checks_file'): 
             # Special logic for Retry Mode to detect "finished" files that are not 100% "total"
             # If a file has NO running workers, but is in active_files, it might be waiting for start
             # OR it might have finished.
             # We rely on 'workers_to_move' logic below? No, that only moves RUNNING workers.
             # We need to detect stopped workers.
             pass
                
        if not active_files:
            logger.info("🎉 All files completed!")
            self.running = False
            return

        # Check if any workers are on completed files
        workers_to_move = [w for w in self.workers if w['file'] in completed_files and w['process'].poll() is None]
        
        if workers_to_move:
            logger.info(f"🔄 Reassigning {len(workers_to_move)} workers from completed files...")
            
            # Stop them
            for w in workers_to_move:
                self.stop_worker(w)
                self.workers.remove(w)
                
            # Restart them on active files
            # Simple strategy: Round robin assignment to active files
            # Note: This is a bit complex because we need to adjust total_workers count for the target file
            # For simplicity in this version, we just append them as "extra" workers
            # They will process a subset based on a new total count?
            # Actually, changing total_workers dynamically is hard because existing workers rely on modulo arithmetic.
            # BETTER STRATEGY:
            # Just start them as NEW workers on the target file with a higher ID.
            # E.g. if file A has 2 workers (1/2, 2/2), adding a 3rd means we start it as...
            # Wait, the modulo logic requires knowing the TOTAL count upfront.
            # If we add a worker, we mess up the division of labor unless we restart ALL workers for that file.
            
            # DECISION: Restart ALL workers for the target file with new counts.
            
            # 1. Pick target file (NEXT one in the list)
            # We want to preserve the order defined in config.EXCEL_FILES
            target_file = None
            for f in self.files:
                if f in active_files:
                    target_file = f
                    break
            
            if not target_file:
                logger.error("Logic error: No target file found despite active_files existing")
                return
            
            # 2. Find all current workers on this target file
            current_workers = [w for w in self.workers if w['file'] == target_file]
            
            # 3. Stop them
            for w in current_workers:
                self.stop_worker(w)
                self.workers.remove(w)
                
            # 4. Calculate new total count
            new_total_count = len(current_workers) + len(workers_to_move)
            
            logger.info(f"⚡ Boosting {Path(target_file).stem} to {new_total_count} workers")
            
            # 5. Start new set of workers
            # We reuse global IDs just for logging, or generate new ones
            next_global_id = max([w['global_id'] for w in self.workers] or [0]) + 1
            
            for i in range(new_total_count):
                self.start_worker(next_global_id + i, target_file, i + 1, new_total_count)
                time.sleep(0.2)  # Reduced delay for faster redistribution

    def monitor(self):
        """Monitor progress loop"""
        start_time = datetime.now()
        
        logger.info("\n📊 Monitoring Progress...")
        
        while self.running:
            time.sleep(5)
            
            # Update session activity from log
            if self.log_handler:
                try:
                    while True:
                        line = self.log_handler.readline()
                        if not line: break
                        if "submitted successfully" in line:
                            self.session_success_count += 1
                except Exception as e:
                    logger.error(f"Error reading log for stats: {e}")

            # Check for redistribution needs
            if not hasattr(self, 'checks_file'):
                self.redistribute_workers()
            if not self.running:
                break
                
            # Display stats
            print(f"\r{'=' * 80}")
            print(f"⏱️  Elapsed: {str(datetime.now() - start_time).split('.')[0]}")
            
            total_processed = 0
            total_checks = 0
            
            hidden_count = 0
            for f in self.files:
                tracker = self.file_trackers[f]
                
                if self.checks_file_set:
                    # Retry Mode: Calculate progress based on retry set
                    # This is heavy but necessary for "correct data"
                    # We need to know how many of self.checks_file_set are in this file's DB
                    completed = 0
                    try:
                        from progress_db import ProgressTracker
                        db_tracker = ProgressTracker(tracker.db_file)
                        all_completed = db_tracker.get_all() # Get set of all completed
                        # Intersection of (All Completed) AND (Retry Target)
                        # Note: This checks GLOBAL progress of the retry set against this file's DB
                        # If the retry set spans multiple files, this logic is tricky.
                        # However, since we are only running 1 file (Sep), we should confirm intersection
                        completed_in_subset = len(all_completed.intersection(self.checks_file_set))
                        
                        # Total target for this file? 
                        # We don't know exactly which of the 7914 checks belong to THIS file without scanning Excel again.
                        # But we can approximate or just show "Completed in Subset"
                        completed = completed_in_subset
                        # For total, we use the size of the retry set (or maybe just keep it simple)
                        total = len(self.checks_file_set) 
                        
                    except Exception as e:
                        logger.error(f"Error calculating filtered progress: {e}")
                        completed = 0
                        total = 0
                else:
                    completed = tracker.get_completed_count()
                    total = tracker.total_checks

                total_processed += completed
                total_checks += total
                
                pct = (completed / total * 100) if total > 0 else 0
                
                # Hide months that are 90%+ completed (Standard Mode)
                if not self.checks_file_set and pct >= HIDE_PROGRESS_THRESHOLD:
                    hidden_count += 1
                    continue
                
                workers_on_file = len([w for w in self.workers if w['file'] == f and w['process'].poll() is None])
                
                if hasattr(self, 'checks_file'):
                    status = f"🔄 RETRYING ({workers_on_file} workers)"
                    # Display for Retry Mode
                    print(f"📄 {tracker.basename[:30]:<30} | {completed:>5} fixed | {status}")
                else:
                    status = "✅ DONE" if completed >= total else f"Running ({workers_on_file} workers)"
                    print(f"📄 {tracker.basename[:30]:<30} | {completed:>5}/{total:<5} ({pct:>5.1f}%) | {status}")
            
            if hidden_count > 0:
                print(f"   ... ({hidden_count} month(s) hidden, ≥{HIDE_PROGRESS_THRESHOLD}% done)")
                
            print(f"{'=' * 80}")
            
            if self.checks_file_set:
                 # Global Retry Progress
                 total_retry_target = len(self.checks_file_set)
                 
                 # Aggregate all DBs
                 global_completed = set()
                 for f in self.files:
                     try:
                         from progress_db import ProgressTracker
                         trk = ProgressTracker(FileProgressTracker(f).db_file)
                         global_completed.update(trk.get_all())
                     except: pass
                 
                 retry_done = len(global_completed.intersection(self.checks_file_set))
                 
                 print(f"Total Retry Progress: {retry_done}/{total_retry_target} ({retry_done/total_retry_target*100:.1f}%)")
                 print(f"Session Activity:     {self.session_success_count} checks processed this run")
            else:
                print(f"Total Progress: {total_processed}/{total_checks} ({total_processed/total_checks*100:.1f}%)")
            
            print(f"{'=' * 80}\n")
            
            # Check if all processes died unexpectedly
            alive = len([w for w in self.workers if w['process'].poll() is None])
            if alive == 0 and self.running:
                # If identifying specific checks, workers might finish quickly if no checks match file
                if hasattr(self, 'checks_file'):
                    logger.info("Accessing file transition logic...")
                    
                    # Mark current files as processed
                    current_files = set(w['file'] for w in self.workers)
                    if not current_files and self.files:
                         # Initial case or weird state
                         pass
                    else:
                        for f in current_files:
                            logger.info(f"Marking {Path(f).stem} as processed (workers finished)")
                            self.processed_files.add(f)
                    
                    # Clear dead worker references
                    self.workers.clear()
                    
                    # Attempt to move to next file
                    self.redistribute_workers() 
                    
                    # Re-check alive status after redistribution attempt
                    alive = len([w for w in self.workers if w['process'].poll() is None])
                    if alive > 0:
                        logger.info("Moved to next file successfully")
                        continue
                    
                    if not self.running:
                        logger.info("All checks processed across all files.")
                        break

                logger.error("❌ All workers died unexpectedly (or finished without transition)!")
                break

    def stop_all(self):
        """Stop all workers and cleanup resources"""
        logger.info("Stopping all workers...")
        for w in self.workers:
            self.stop_worker(w)
        self.workers.clear()  # Clear worker list
        logger.info("All workers stopped")

def refresh_session():
    """Refresh session cookies ONCE before starting workers"""
    logger.info("=" * 80)
    logger.info("🔄 STEP 1: Refreshing Session Cookies")
    logger.info("=" * 80)
    
    # Check if refresh_cookies.py exists
    refresh_script = Path("refresh_cookies.py")
    if not refresh_script.exists():
        logger.warning("⚠️  refresh_cookies.py not found")
        input("\nPress ENTER when cookies are ready in config.py: ")
        return True
    
    try:
        # Start Chrome in debug mode
        logger.info("Starting Chrome...")
        subprocess.Popen(
            ["START_CHROME_DEBUG.bat"],
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )
        
        time.sleep(3)
        input("\n👉 Press ENTER after you've logged in: ")
        
        # Extract cookies
        logger.info("🔄 Extracting cookies...")
        result = subprocess.run(
            [sys.executable, "refresh_cookies.py"],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='ignore',
            timeout=120  # Increased from 30s to prevent timeouts
        )
        
        if result.returncode == 0:
            logger.info("✅ Cookies extracted")
            return True
        else:
            logger.error(f"❌ Cookie extraction failed: {result.stderr}")
            return False
            
    except Exception as e:
        logger.error(f"Error during session refresh: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description='Master Controller')
    parser.add_argument('--workers', type=int, default=8, help='Total number of workers')
    parser.add_argument('--dry-run', action='store_true', help='Dry run mode')
    parser.add_argument('--checks-file', type=str, help='JSON file containing specific check numbers to process')
    args = parser.parse_args()

    if not refresh_session():
        sys.exit(1)
        
    manager = WorkerManager(total_workers=args.workers)
    if args.checks_file:
        manager.checks_file = args.checks_file
        # Load checks set for progress monitoring
        try:
            with open(args.checks_file, 'r', encoding='utf-8') as f:
                manager.checks_file_set = set(json.load(f))
            logger.info(f"Loaded {len(manager.checks_file_set)} checks to monitor for progress")
        except Exception as e:
            logger.error(f"Failed to load checks file for monitoring: {e}")
            
        # Disable hiding completed months when running specific checks
        global HIDE_PROGRESS_THRESHOLD
        HIDE_PROGRESS_THRESHOLD = 101
    
    # Check worker/key ratio
    num_keys = len(config.GEMINI_API_KEYS)
    if num_keys > 0:
        ratio = args.workers / num_keys
        if ratio > 2:
            logger.warning("=" * 80)
            logger.warning(f"⚠️  WARNING: High Worker/Key Ratio ({ratio:.1f} workers per key)")
            logger.warning(f"   You have {args.workers} workers but only {num_keys} active Gemini API keys.")
            logger.warning("   This may cause Rate Limiting (429) or Suspension.")
            logger.warning("   RECOMMENDED: Add more keys to config.py or reduce workers.")
            logger.warning("=" * 80)
            time.sleep(3)  # Give user time to read
            
    try:
        manager.start_initial_distribution()
        manager.monitor()
    except KeyboardInterrupt:
        logger.info("Interrupted!")
        manager.stop_all()

if __name__ == "__main__":
    main()
