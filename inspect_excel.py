"""
Quick script to inspect Excel file structure
"""
try:
    from openpyxl import load_workbook
    
    excel_file = r"c:\Users\User\Desktop\Новая папка (2)\Асилбекова_Октябрь_тахрирлаш_14626_та.xlsx"
    
    print("Opening Excel file...")
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"\nTotal columns: {len(headers)}")
    print("\nColumn headers:")
    for idx, header in enumerate(headers):
        print(f"  [{idx}] {header}")
    
    # Show first 3 data rows
    print("\n\nFirst 3 data rows:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), start=2):
        print(f"\nRow {row_idx}:")
        for idx, value in enumerate(row[:15]):  # Show first 15 columns
            if value is not None:
                print(f"  [{idx}] {headers[idx]}: {value}")
    
    wb.close()
    print("\n\nDone!")
    
except ImportError:
    print("Error: openpyxl not installed")
    print("Run: pip install openpyxl")
except Exception as e:
    print(f"Error: {e}")
