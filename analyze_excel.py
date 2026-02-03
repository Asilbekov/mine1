"""
Excel Data Analyzer - Creates Markdown Report
This script analyzes the Excel file and creates a detailed markdown report
of its structure to help optimize the automation script.
"""

def analyze_excel():
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("=" * 80)
        print("ERROR: openpyxl not installed")
        print("=" * 80)
        print("\nPlease install Python first, then run:")
        print("  pip install openpyxl")
        print("\nPython installation:")
        print("  1. Download from: https://www.python.org/downloads/")
        print("  2. Run installer")
        print("  3. CHECK 'Add Python to PATH'")
        print("  4. Complete installation")
        print("  5. Open new Command Prompt")
        print("  6. Run: pip install openpyxl")
        print("=" * 80)
        return
    
    excel_file = r"c:\Users\User\Desktop\Новая папка (2)\Асилбекова_Октябрь_тахрирлаш_14626_та.xlsx"
    output_file = r"c:\Users\User\Desktop\Новая папка (2)\excel_structure.md"
    
    print("Analyzing Excel file...")
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    
    # Get sample data
    sample_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
        sample_rows.append((row_idx, row))
    
    # Count total rows
    total_rows = ws.max_row - 1  # Excluding header
    
    # Analyze data types and uniqueness
    column_analysis = {}
    for idx, header in enumerate(headers):
        values = set()
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(102, ws.max_row), values_only=True), start=2):
            if idx < len(row) and row[idx] is not None:
                values.add(str(row[idx]))
        
        column_analysis[idx] = {
            'header': header,
            'unique_count': len(values),
            'sample_values': list(values)[:5]
        }
    
    # Create markdown report
    md_content = f"""# Excel File Structure Analysis

## File Information
- **File**: `Асилбекова_Октябрь_тахрирлаш_14626_та.xlsx`
- **Total Rows**: {total_rows:,} (excluding header)
- **Total Columns**: {len(headers)}

## Column Structure

| Index | Column Name | Unique Values | Sample Data |
|-------|-------------|---------------|-------------|
"""
    
    for idx, analysis in column_analysis.items():
        header = analysis['header'] or f'Column_{idx}'
        unique = analysis['unique_count']
        samples = ', '.join([str(v)[:30] for v in analysis['sample_values'][:3]])
        md_content += f"| {idx} | {header} | {unique} | {samples} |\n"
    
    md_content += f"""

## Sample Data (First 5 Rows)

"""
    
    for row_num, row_data in sample_rows:
        md_content += f"### Row {row_num}\n\n"
        for idx, value in enumerate(row_data[:20]):  # First 20 columns
            if value is not None:
                header = headers[idx] if idx < len(headers) else f'Column_{idx}'
                md_content += f"- **{header}**: {value}\n"
        md_content += "\n"
    
    md_content += """
## Key Findings for Automation

### Check Number Column
"""
    
    # Find check number column
    check_col_idx = None
    for idx, header in enumerate(headers):
        if header and ('Чек' in str(header) or 'рақами' in str(header)):
            check_col_idx = idx
            md_content += f"- **Found at index {idx}**: `{header}`\n"
            md_content += f"- **Unique values**: {column_analysis[idx]['unique_count']:,}\n"
            md_content += f"- **Sample**: {', '.join([str(v) for v in column_analysis[idx]['sample_values'][:5]])}\n"
            break
    
    if check_col_idx is None:
        md_content += "- **WARNING**: Could not find check number column!\n"
    
    md_content += """
### Other Important Columns

Look for columns that might contain:
- **Terminal ID**: Terminal identifier
- **TIN**: Tax identification number  
- **Payment Date**: Transaction date
- **Commission TIN**: Commission entity TIN
- **Product Code**: MXIK product code

### Recommendations for Dynamic Extraction

Based on the analysis above, update `automate_checks.py` to:

1. **If unique count = 1**: Use as constant (all checks have same value)
2. **If unique count > 1**: Extract dynamically from Excel column
3. **If unique count ≈ total rows**: Use as unique identifier (like check number)

### Columns to Extract

"""
    
    # Make recommendations
    for idx, analysis in column_analysis.items():
        header = analysis['header']
        unique = analysis['unique_count']
        
        if unique == 1:
            md_content += f"- **{header}** (Index {idx}): CONSTANT - Use config default\n"
        elif unique > 1 and unique < total_rows * 0.1:
            md_content += f"- **{header}** (Index {idx}): FEW UNIQUE - May be Terminal ID or similar\n"
        elif unique >= total_rows * 0.9:
            md_content += f"- **{header}** (Index {idx}): MOSTLY UNIQUE - Likely check number or transaction ID\n"
    
    # Write to file
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(md_content)
    
    wb.close()
    
    print(f"\n✓ Analysis complete!")
    print(f"✓ Report saved to: {output_file}")
    print(f"\nTotal checks in Excel: {total_rows:,}")
    print(f"\nPlease review the markdown file to see the structure.")
    
    # Also print to console
    print("\n" + "=" * 80)
    print("QUICK SUMMARY")
    print("=" * 80)
    print(md_content[:2000] + "\n...\n(See full report in excel_structure.md)")

if __name__ == "__main__":
    analyze_excel()
