"""
Excel Parser - Load checks from uploaded Excel file
"""

import json
import os
import io
from http.server import BaseHTTPRequestHandler
import psycopg2
from psycopg2.extras import RealDictCursor

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

DATABASE_URL = os.environ.get('DATABASE_URL', '')

def get_db_connection():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

def get_config():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT key, value FROM "ProjectConfig"')
    rows = cur.fetchall()
    conn.close()
    return {row['key']: row['value'] for row in rows}

def add_log(level, message, details=None):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            '''INSERT INTO "AutomationLog" (id, timestamp, level, message, details)
               VALUES (gen_random_uuid(), NOW(), %s, %s, %s)''',
            (level, message, json.dumps(details) if details else None)
        )
        conn.commit()
        conn.close()
    except:
        pass

def parse_excel_from_db():
    """Parse Excel file from database and create Check records"""
    
    if not load_workbook:
        return {'success': False, 'error': 'openpyxl not installed'}
    
    config = get_config()
    
    # Get Excel file from database
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        '''SELECT id, data, "originalName" FROM "UploadedFile" 
           WHERE type = 'excel' AND "isActive" = true 
           ORDER BY "createdAt" DESC LIMIT 1'''
    )
    file_row = cur.fetchone()
    
    if not file_row or not file_row.get('data'):
        conn.close()
        return {'success': False, 'error': 'No Excel file found in database'}
    
    file_name = file_row.get('originalName', 'unknown')
    file_id = file_row.get('id')
    add_log('INFO', f'Parsing Excel file: {file_name}')
    
    try:
        # Load workbook from bytes
        file_bytes = bytes(file_row['data'])
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append(f'col_{len(headers)}')
        
        # Find column indices
        check_col = config.get('CHECK_NUMBER_COLUMN', 'receipt_id').lower()
        check_idx = None
        terminal_idx = None
        tin_idx = None
        date_idx = None
        
        for i, h in enumerate(headers):
            h_lower = h.lower()
            if check_col in h_lower or 'receipt' in h_lower or 'check' in h_lower:
                check_idx = i
            if 'terminal' in h_lower:
                terminal_idx = i
            if 'tin' in h_lower:
                tin_idx = i
            if 'date' in h_lower or 'дата' in h_lower:
                date_idx = i
        
        if check_idx is None:
            # Try first column
            check_idx = 0
        
        add_log('INFO', f'Found columns - check: {check_idx}, terminal: {terminal_idx}, tin: {tin_idx}, date: {date_idx}')
        
        # Parse rows
        checks_data = []
        row_num = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or not row[check_idx]:
                continue
            
            receipt_id = str(row[check_idx]).strip()
            if not receipt_id or receipt_id == 'None':
                continue
            
            terminal_id = str(row[terminal_idx]).strip() if terminal_idx is not None and row[terminal_idx] else config.get('DEFAULT_TERMINAL_ID', 'EP000000000551')
            tin = str(row[tin_idx]).strip() if tin_idx is not None and row[tin_idx] else config.get('DEFAULT_TIN', '')
            payment_date = str(row[date_idx]).strip() if date_idx is not None and row[date_idx] else None
            
            # Build payment_id
            check_number_part = receipt_id.zfill(16)
            payment_id = f"{terminal_id}{check_number_part}"
            
            checks_data.append({
                'receipt_id': receipt_id,
                'payment_id': payment_id,
                'terminal_id': terminal_id,
                'tin': tin,
                'payment_date': payment_date,
                'file_id': file_id
            })
        
        wb.close()
        
        if not checks_data:
            conn.close()
            return {'success': False, 'error': 'No valid check data found in Excel'}
        
        add_log('INFO', f'Found {len(checks_data)} checks in Excel')
        
        # Check for existing checks to avoid duplicates
        cur.execute('SELECT "receiptId" FROM "Check"')
        existing = set(row['receiptId'] for row in cur.fetchall())
        
        new_checks = [c for c in checks_data if c['receipt_id'] not in existing]
        
        if not new_checks:
            conn.close()
            return {'success': True, 'message': 'All checks already exist', 'total': len(checks_data), 'new': 0}
        
        # Insert new checks
        for check in new_checks:
            cur.execute(
                '''INSERT INTO "Check" (id, "receiptId", "paymentId", "terminalId", tin, "paymentDate", status, "createdAt", "updatedAt", "fileId")
                   VALUES (gen_random_uuid(), %s, %s, %s, %s, %s, 'pending', NOW(), NOW(), %s)''',
                (check['receipt_id'], check['payment_id'], check['terminal_id'], check['tin'], check['payment_date'], check['file_id'])
            )
        
        conn.commit()
        conn.close()
        
        add_log('INFO', f'Created {len(new_checks)} new check records')
        
        return {
            'success': True,
            'total': len(checks_data),
            'new': len(new_checks),
            'skipped': len(checks_data) - len(new_checks)
        }
        
    except Exception as e:
        add_log('ERROR', f'Excel parse error: {str(e)}')
        return {'success': False, 'error': str(e)}


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        """Parse Excel and create check records"""
        try:
            result = parse_excel_from_db()
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(result).encode())
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
